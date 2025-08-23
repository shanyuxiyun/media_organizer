import os
import argparse
import openpyxl
from openpyxl import Workbook

def merge_excel_files(directory, output_filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Files"
    ws.append(['Path', 'Size', 'Hash', 'HashSize'])

    for i in os.listdir(directory):
        if i.endswith('.xlsx'):
            s = openpyxl.load_workbook(os.path.join(directory, i))['Files']
            for row in s.iter_rows(min_row=2):
                ws.append([cell.value for cell in row])

    wb.save(os.path.join(directory, output_filename))

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Merge Excel files.')
    parser.add_argument('-d', '--directory', type=str, required=True, help='The directory containing the Excel files to merge.')
    parser.add_argument('-o', '--output', type=str, default='Merge.xlsx', help='The name of the output file.')
    args = parser.parse_args()

    merge_excel_files(args.directory, args.output)